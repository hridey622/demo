import json
import re
import sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
from anthropic import Anthropic
from openpyxl.worksheet.datavalidation import DataValidation
import whisper

try:
    from yt_dlp import YoutubeDL
    HAS_YT_DLP_MODULE = True
except ImportError:
    HAS_YT_DLP_MODULE = False

# Define all valid field values - all fields (except language) have multiple choice values
FIELD_VALUES = {
    "tut":["0","1"],
    "language": ["English", "Spanish", "Hindi", "Mandarin", "Portuguese", "Sanskrit", "Other", "Mixed"],
    "primary_modality": ["vocal guided", "vocal chanting", "instrumental only", "visual only", "mixed"],
    "language_presence": ["continuous speech", "intermittent speech", "chant mantra", "none"],
    "therapeutic_explicitness": ["implicit", "experimental", "explicit", "other"],
    "session_structure": ["clearly staged", "partially structured", "unstructured"],
    "pacing_adequacy": ["rushed", "slow stable", "fluctuating", "moderate"],
    "source_lineage": ["classical textual", "oral tradition", "modern traditional synthesis", "non traditional"],
    "symbolic_density": ["high", "medium", "low"],
    "modern_intrusion_level": ["none", "minimal", "moderate", "high"],
    "emotional_valence": ["calming", "grounding", "uplifting", "devotional", "introspective", "mixed"],
    "affective_stability": ["highly_stable", "mostly_stable", "fluctuating", "unstable"],
    "arousal_trajectory": ["decreasing", "steady_low", "steady_moderate", "increasing"],
    "instruction_density": ["none", "sparse", "moderate", "dense"],
    "cognitive_demand": ["passive_receptivity", "light_following", "active_engagement", "effortful_reasoning"],
    "inner_silence_support": ["strong", "moderate", "weak"],
    "directive_strength": ["invitational", "neutral", "commanding", "coercive"],
    "dependency_risk": ["none", "low", "moderate", "high"],
    "claim_severity": ["no claims", "general wellbeing", "quasi medical", "medical cure"],
    "autonomic_orientation": ["parasympathetic", "balancing", "sympathetic activating"],
    "best_fit_states": ["high anxiety", "fatigue", "restlessness", "emotional heaviness", "post rest reintegration"],
    "sleep_proximity_safety": ["safe", "caution", "unsafe"],
    "cultural_accessibility": ["universal", "requires familiarity", "tradition heavy"],
    "locale fit": ["metro", "town", "rural"],
    "environmental_compatibility": ["home quiet", "shared household", "outdoor"],
    "sensory_purity": ["clean", "mildly layered", "overstimulating"],
    "rhythmic_stability": ["highly stable", "gently evolving", "irregular"],
    "visual_motion_profile": ["static", "slow natural motion", "fast artificial"],
    "repetitiveness_quality": ["supportive", "neutral", "hypnotic risk"],
    "attention_capture_style": ["soft holding", "gentle anchoring", "gripping"],
    "internal_consistency": ["coherent", "minor drift", "inconsistent"],
    "production_intrusion": ["minimal", "noticeable", "distracting"],
    "commercial_influence": ["none", "subtle", "explicit"],
    "therapeutic_suitability_score": ["1", "2", "3", "4", "5"],
    "Risk Flag Summary": "tell if there is any risk watching this video and tell what kind of risk",
    "One-line Neutral Description" : "give a one line description of the video"

}
# Only fields with meaningful quality/therapeutic order get scored
# Higher number = generally more desirable for therapeutic/guided meditation context
SCORE_MAP = {
    "session_structure": {
        "clearly_staged": 5,
        "partially_structured": 3,
        "unstructured": 2,
    },
    "pacing_adequacy": {
        "slow_stable": 5,
        "moderate": 4,
        "fluctuating": 2,
        "rushed": 1,
    },
    "symbolic_density": {
        "low": 4,       # usually cleaner for beginners / less cognitive load
        "medium": 3,
        "high": 2,
    },
    "modern_intrusion_level": {
        "none": 5,
        "minimal": 4,
        "moderate": 2,
        "high": 1,
    },
    "affective_stability": {
        "highly_stable": 5,
        "mostly_stable": 4,
        "fluctuating": 2,
        "unstable": 1,
    },
    "arousal_trajectory": {
        "decreasing": 5,          # down-regulation is often desired
        "steady_low": 4,
        "steady_moderate": 3,
        "increasing": 1,
    },
    "inner_silence_support": {
        "strong": 5,
        "moderate": 3,
        "weak": 2,
    },
    "directive_strength": {
        "invitational": 5,
        "neutral": 4,
        "commanding": 2.5,
        "coercive": 1,
    },
    "dependency_risk": {
        "none": 5,
        "low": 4,
        "moderate": 2.5,
        "high": 1,
    },
    "claim_severity": {
        "no_claims": 5,
        "general_wellbeing": 4,
        "quasi_medical": 2,
        "medical_cure": 1,
    },
    "sleep_proximity_safety": {
        "safe": 5,
        "caution": 3,
        "unsafe": 1,
    },
    "sensory_purity": {
        "clean": 5,
        "mildly_layered": 3,
        "overstimulating": 1,
    },
    "rhythmic_stability": {
        "highly_stable": 5,
        "gently_evolving": 4,
        "irregular": 2,
    },
    "repetitiveness_quality": {
        "supportive": 5,
        "neutral": 3,
        "hypnotic_risk": 1,
    },
    "internal_consistency": {
        "coherent": 5,
        "minor_drift": 3,
        "inconsistent": 1,
    },
    "production_intrusion": {
        "minimal": 5,
        "noticeable": 3,
        "distracting": 1,
    },
    "commercial_influence": {
        "none": 5,
        "subtle": 3,
        "explicit": 1,
    },
    # You can add more later — but avoid scoring categorical fields without clear order
    # e.g. do NOT score: primary_modality, emotional_valence, best_fit_states, language, etc.
}
# Fields that can have multiple values selected (array output)
MULTI_VALUE_FIELDS = ["best_fit_states"]

from groq import Groq
from pydub import AudioSegment
from pydub.silence import detect_silence
import tempfile

# Recommended: load from environment
import os

import tempfile
import os
import contextlib
from pathlib import Path

# ────────────────────────────────────────────────
#   Centralized audio downloader + cleanup
# ────────────────────────────────────────────────

@contextlib.contextmanager
def download_audio_temp(video_url: str, preferred_format="mp3"):
    """
    Context manager: downloads audio → yields path → deletes after use
    """
    temp_dir = tempfile.mkdtemp(prefix="yt_audio_")
    base_path = os.path.join(temp_dir, "audio")

    ydl_opts = {
        "format": "bestaudio/best",
        "outtmpl": f"{base_path}.%(ext)s",
        "quiet": True,
        "no_warnings": True,
        "postprocessors": [{
            "key": "FFmpegExtractAudio",
            "preferredcodec": preferred_format,
            "preferredquality": "128",
        }],
    }

    audio_path = None

    try:
        with YoutubeDL(ydl_opts) as ydl:
            info = ydl.extract_info(video_url, download=True)
            # After post-processing we should have .mp3
            audio_path = f"{base_path}.{preferred_format}"

            if not os.path.exists(audio_path):
                # Fallback: maybe postprocessor failed → try original
                audio_path = f"{base_path}.webm"
                if not os.path.exists(audio_path):
                    audio_path = None

        if audio_path is None:
            raise FileNotFoundError("Could not find downloaded audio file")

        yield audio_path, info.get('duration', 0)

    finally:
        # Aggressive cleanup
        if audio_path and os.path.exists(audio_path):
            try:
                os.remove(audio_path)
            except:
                pass

        try:
            os.rmdir(temp_dir)
        except:
            pass
client = Groq(api_key=os.getenv("GROQ_API_KEY"))
def analyze_audio_pacing(audio_path, duration):
    print("  [PACING] Analyzing audio...")

    if duration == 0:
        duration = len(AudioSegment.from_file(audio_path)) / 1000

    print(f"  [PACING] Analyzing {duration:.0f}s audio...")

    try:
        audio = AudioSegment.from_file(audio_path)
    except Exception as e:
        print(f"  [PACING] Cannot load audio: {e}")
        return None

    silences = detect_silence(
        audio,
        min_silence_len=500,
        silence_thresh=audio.dBFS - 16
    )

    # ── rest of your pacing logic stays almost the same ──
    total_duration_s = duration
    pause_durations = [(end - start)/1000 for start, end in silences]
    num_pauses = len(pause_durations)

    if num_pauses == 0:
        return "rushed"

    avg_pause = sum(pause_durations) / num_pauses
    total_silence = sum(pause_durations)
    silence_ratio = total_silence / total_duration_s
    variance = sum((p - avg_pause)**2 for p in pause_durations) / num_pauses if num_pauses > 1 else 0
    pauses_per_min = num_pauses / (total_duration_s / 60)

    if avg_pause < 1.0 and pauses_per_min > 12:
        return "rushed"
    elif avg_pause > 3.0 and pauses_per_min < 4 and silence_ratio > 0.15:
        return "slow stable"
    elif variance > 4.0 or (max(pause_durations) > 5 * min(pause_durations) if pause_durations else False):
        return "fluctuating"
    else:
        return "moderate"
# If you want to hard-code (not recommended for shared code):
# client = Groq(api_key="gsk_xxxxx")
def extract_video_id(url):
    """Extract video ID from YouTube URL."""
    patterns = [r'(?:youtube\.com\/watch\?v=|youtu\.be\/|youtube\.com\/embed\/)([^&\n?#]+)']
    for pattern in patterns:
        match = re.search(pattern, url)
        if match:
            return match.group(1)
    return None
# def transcribe_any_file(file_path): # using whisper
#     # 1. Load the Model
#     # "base" is a good middle ground (about 140MB). 
#     # Use "tiny" for speed or "small/medium/large" for better accuracy.
#     model = whisper.load_model("base")

#     # 2. Transcribe
#     # fp16=False is CRITICAL for Windows/CPU users to prevent crashing.
#     result = model.transcribe(file_path, fp16=False)
    
#     # 3. Extract the Text
#     return result["text"]
def get_transcript(video_id: str, audio_path: str, duration: int):
    video_url = f"https://www.youtube.com/watch?v={video_id}"

    # First try → subtitles (fast, no download)
    try:
        ydl_opts = {
            'quiet': True,
            'no_warnings': True,
            'skip_download': True,
            'writesubtitles': True,
            'writeautomaticsub': True,
            'subtitleslangs': ['en'],
        }
        with YoutubeDL(ydl_opts) as ydl:
            info = ydl.extract_info(video_url, download=False)

        for source in ['subtitles', 'automatic_captions']:
            if source in info and 'en' in info[source]:
                for sub in info[source]['en']:
                    try:
                        import urllib.request
                        with urllib.request.urlopen(sub['url']) as resp:
                            vtt = resp.read().decode('utf-8', errors='replace')
                            text = parse_vtt(vtt)
                            if text.strip():
                                print("  ✓ Got embedded/auto subtitles")
                                return text
                    except:
                        continue
    except Exception as e:
        print(f"  [subtitles] failed: {e}")

    # Fallback → download audio + Groq Whisper
    print("  [transcript] No subtitles → using audio + Whisper...")


    try:
        with open(audio_path, "rb") as f:
            transcription = client.audio.transcriptions.create(
                file=(os.path.basename(audio_path), f.read()),
                model="whisper-large-v3-turbo",
                response_format="verbose_json",
                language="en"
            )
        text = transcription.text.strip()
        if text:
            print("  ✓ Whisper transcription successful")
            return text
    except Exception as e:
        print(f"  [Whisper] failed: {e}")
        return None

    return None

def parse_vtt(vtt_content):
    """Parse VTT caption format."""
    lines = vtt_content.split('\n')
    transcript = []
    for line in lines:
        if line.startswith('WEBVTT') or '-->' in line or not line.strip():
            continue
        text = re.sub(r'<[^>]+>', '', line.strip())
        if text:
            transcript.append(text)
    return ' '.join(transcript)

def get_video_metadata(video_id):
    """Extract video metadata."""
    try:
        if not HAS_YT_DLP_MODULE:
            return None
        
        ydl_opts = {'quiet': True, 'no_warnings': True, 'skip_download': True}
        
        with YoutubeDL(ydl_opts) as ydl:
            data = ydl.extract_info(f'https://www.youtube.com/watch?v={video_id}', download=False)
        
        return {
            'title': data.get('title', 'Unknown'),
            'channel': data.get('uploader', 'Unknown'),
            'duration': data.get('duration', 0),
            'description': data.get('description', ''),
            'tags': data.get('tags', []),
        }
    except Exception as e:
        return None
def analyze_video_with_groq(url, transcript, metadata):
    """Use Groq to analyze video and return structured field values."""
    
    if not transcript and not metadata:
        return {'error': 'Could not extract transcript or metadata'}
    
    metadata_text = json.dumps(metadata, indent=2) if metadata else "Metadata unavailable"
    transcript_text = transcript[:8000] + "...[truncated]" if transcript and len(transcript) > 8000 else (transcript or "Transcript unavailable")
    
    # Step 1: Classify video type first
    classification_prompt = f"""Analyze this YouTube video and classify its type.

VIDEO URL: {url}

METADATA:
{metadata_text}

TRANSCRIPT (first part):
{transcript_text[:3000]}

---

Classify this video into ONE of the following categories:
1. "therapy" - A therapeutic video meant for relaxation, meditation, healing, guided practice, wellness, yoga nidra, sound healing, breathwork,chanting, or similar therapeutic/wellness content that a user would EXPERIENCE or PRACTICE along with.
2. "tutorial" - An educational/instructional video that TEACHES or EXPLAINS concepts, techniques, or information. This includes lectures, how-to guides, educational content, talks, interviews, discussions, or any content primarily meant to educate or inform rather than be practiced along with.

IMPORTANT DISTINCTIONS:
- A "guided meditation" where you follow along = THERAPY
- A video "explaining how meditation works" or "teaching meditation techniques" = TUTORIAL
- A "relaxing music/sound bath" to experience = THERAPY  
- A video "about the benefits of sound healing" = TUTORIAL
- A "yoga nidra session" to practice = THERAPY
- A "lecture on yoga philosophy" = TUTORIAL

Return ONLY valid JSON with no extra text:
{{
  "video_type": "therapy" or "tutorial",
  "confidence": "high" or "medium" or "low",
  "reasoning": "brief one-line explanation"
}}
"""
    
    try:
        # First, classify the video
        classification_response = client.chat.completions.create(
            model="llama-3.3-70b-versatile",
            messages=[{"role": "user", "content": classification_prompt}],
            temperature=0.1,
            max_tokens=200,
            response_format={"type": "json_object"}
        )
        
        classification_text = classification_response.choices[0].message.content
        classification_text = re.sub(r'^```json\s*|\s*```$', '', classification_text.strip())
        classification = json.loads(classification_text)
        
        video_type = classification.get("video_type", "").lower()
        reasoning = classification.get("reasoning", "")
        
        if video_type == "tutorial":
            return {
                'error': f'Not a therapy video - classified as tutorial/educational content',
                'classification_reason': reasoning,
                'video_url': url
            }
        
        print(f"  ✓ Classified as therapy video: {reasoning}")
        
    except Exception as e:
        print(f"  ⚠️ Classification failed, proceeding with analysis: {str(e)}")
    
    # Step 2: If therapy video, proceed with full analysis
    valid_values_text = json.dumps(FIELD_VALUES, indent=2)
    
    analysis_prompt = f"""Analyze this YouTube video and extract structured feature values.

VIDEO URL: {url}

METADATA:
{metadata_text}

TRANSCRIPT:
{transcript_text}

---

VALID VALUES FOR EACH FIELD (only use these values exactly as listed):
{valid_values_text}

CRITICAL RULES:
- Return ONLY valid JSON — no extra text, no markdown, no explanation, no ```json block

-Below are the given definitions for the primary_modality field, so use the suitable one based on their description:
-vocal-guided: Spoken language used mainly for explanation, instruction, or narration.
-vocal-chanting: Repetitive, rhythmic, or mantra-like vocal sounds (e.g., “OM”, humming, chanting) with little or no explanatory speech.
-instrumental only: No voice; only music or instrumental sounds.
-visual only: No audio content is implied.
-mixed: Both spoken guidance and chanting/mantra-style vocalizations are present.
-----
-Below are the given definitions for the language_presence field, so use the suitable one based on their description:
-continuous speech: Spoken language is present for most of the content with no long breaks or non-speech segments.
-intermittent speech: Spoken language appears in segments and is interrupted by silence, pauses, or non-speech vocalizations.
-chant/mantra: The content consists mainly of repetitive, rhythmic, or mantra-like vocalizations with little or no explanatory speech.
-none: No speech or vocalization is implied by the transcript.
------
-Below are the given definitions for the symbolic_density field, so use the suitable one based on their description:
-high: Language is highly symbolic, metaphorical, ritualistic, or abstract, with meaning compressed into symbols, chants, poetry, or sacred references and little literal explanation.
-medium: Symbolic or ritual language is present but is explained or supported by clear, literal, instructional, or descriptive speech.
-low: Language is mostly literal, procedural, or factual, with little to no symbolic, metaphorical, or ritual content.
------
-Below are the given definitions for the visual_motion_profile field, so use the suitable one based on the transcript:
-static: Choose this when the transcript indicates continuous chanting, mantra repetition, or minimal/no spoken guidance - suggests the video has still or unchanging visuals.
-slow natural motion: Choose this as the DEFAULT for most therapy/meditation videos. Indicates gentle, calming visual movement like nature scenes, flowing water, clouds, or slow transitions. Use this when there is guided meditation, relaxation instructions, or mixed content.
-fast artificial: Choose this ONLY when the transcript suggests rapid, chaotic, or energetic content - very rare for therapy videos.
------
-Below are the given definitions for directive_strength field, so use the suitable one based on the description:
-invitational: The content uses gentle suggestions, invitations, or optional guidance.
-neutral: The content provides information without strong direction either way.
-commanding: The content gives clear, direct instructions or commands.
-coercive: The content uses strong pressure, manipulation, or forceful directives.
------
-Below are the given definitions for production_intrusion field, so use the suitable one based on the description:
-minimal: Seamless production. No visible production equipment, no distracting logos, no "calls to action" (CTAs) like "like/subscribe" during the content. Focus is entirely on immersion.
-noticeable: Professional but visible production. Clean editing, small persistent logos, or CTAs limited to the very beginning or end of the video.
-distracting: Highly intrusive. Flashy transitions, frequent text overlays, loud music that masks speech, mid-video pop-ups, visible equipment, or poor audio/video quality (background noise/hiss).
------
- For all fields EXCEPT "best_fit_states": return a SINGLE string value from the list
- For "best_fit_states": return an ARRAY of zero or more applicable values from the list
- For "language": pick ONE from the list
- "Risk Flag Summary": short summary — if no risk write "No significant risks identified"
- "One-line Neutral Description": exactly one neutral descriptive sentence
- Do NOT include "therapeutic_suitability_score" — it will be calculated later

Return ONLY this JSON structure:

{{
  "tut": "",
  "video_url": "{url}",
  "language": "",
  "primary_modality": "",
  "language_presence": "",
  "therapeutic_explicitness": "",
  "session_structure": "",
  "pacing_adequacy": "",
  "source_lineage": "",
  "symbolic_density": "",
  "modern_intrusion_level": "",
  "emotional_valence": "",
  "affective_stability": "",
  "arousal_trajectory": "",
  "instruction_density": "",
  "cognitive_demand": "",
  "inner_silence_support": "",
  "directive_strength": "",
  "dependency_risk": "",
  "claim_severity": "",
  "autonomic_orientation": "",
  "best_fit_states": [],
  "sleep_proximity_safety": "",
  "cultural_accessibility": "",
  "locale_fit": "",
  "environmental_compatibility": "",
  "sensory_purity": "",
  "rhythmic_stability": "",
  "visual_motion_profile": "",
  "repetitiveness_quality": "",
  "attention_capture_style": "",
  "internal_consistency": "",
  "production_intrusion": "",
  "commercial_influence": "",
  "Risk Flag Summary": "",
  "One-line Neutral Description": ""
}}
"""
    try:
        response = client.chat.completions.create(
            model="llama-3.3-70b-versatile",          # ← good balance of quality & speed
            #model="llama-3.1-8b-instant",           # ← fastest & cheapest
            messages=[{"role": "user", "content": analysis_prompt}],
            temperature=0.1,                          # low = more deterministic / better JSON
            max_tokens=2000,
            response_format={"type": "json_object"}   # helps enforce JSON (supported on many Groq models)
        )
        
        response_text = response.choices[0].message.content
        
        # Clean up if wrapped in ```json ... ```
        response_text = re.sub(r'^```json\s*|\s*```$', '', response_text.strip())
        
        analysis = json.loads(response_text)
        return analysis
    
    except Exception as e:
        return {'error': f'Groq API error: {str(e)}'}

def create_excel_file(analyses, filename='wellness_video_analysis.xlsx'):
    """Create Excel file with dropdown (data validation) controls for all fields."""

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Video Analysis"

    # -----------------------------
    # STYLES
    # -----------------------------
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    data_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    # -----------------------------
    # FIELD ORDER
    # -----------------------------
    field_names = [
        "tut",
        "video_url", "language", "primary_modality", "language_presence",
        "therapeutic_explicitness", "session_structure", "pacing_adequacy",
        "source_lineage", "symbolic_density", "modern_intrusion_level",
        "emotional_valence", "affective_stability", "arousal_trajectory",
        "instruction_density", "cognitive_demand", "inner_silence_support",
        "directive_strength", "dependency_risk", "claim_severity",
        "autonomic_orientation", "best_fit_states", "sleep_proximity_safety",
        "cultural_accessibility", "locale_fit", "environmental_compatibility",
        "sensory_purity", "rhythmic_stability", "visual_motion_profile",
        "repetitiveness_quality", "attention_capture_style", "internal_consistency",
        "production_intrusion", "commercial_influence",
        "therapeutic_suitability_score", "Risk Flag Summary", "One-line Neutral Description"
    ]

    # -----------------------------
    # CREATE HIDDEN LIST SHEET
    # -----------------------------
    list_ws = wb.create_sheet("__lists__")
    list_ws.sheet_state = "hidden"

    list_col_map = {}
    col_idx = 1

    for field, values in FIELD_VALUES.items():
        list_col_map[field] = col_idx
        for row_idx, val in enumerate(values, start=1):
            list_ws.cell(row=row_idx, column=col_idx, value=val)
        col_idx += 1

    # -----------------------------
    # WRITE HEADERS
    # -----------------------------
    for col, field in enumerate(field_names, start=1):
        cell = ws.cell(row=1, column=col, value=field)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # -----------------------------
    # ADD DROPDOWNS
    # -----------------------------
    MAX_ROWS = max(1000, len(analyses) + 10)

    for col, field in enumerate(field_names, start=1):
        if field not in FIELD_VALUES:
            continue

        list_col = list_col_map[field]
        values_len = len(FIELD_VALUES[field])

        formula = (
            f"=__lists__!"
            f"${get_column_letter(list_col)}$1:"
            f"${get_column_letter(list_col)}${values_len}"
        )

        dv = DataValidation(
            type="list",
            formula1=formula,
            allow_blank=True,
            showDropDown=True
        )

        ws.add_data_validation(dv)
        dv.add(
            f"{get_column_letter(col)}2:"
            f"{get_column_letter(col)}{MAX_ROWS}"
        )

    # -----------------------------
    # WRITE DATA
    # -----------------------------
    for row_idx, analysis in enumerate(analyses, start=2):
        for col_idx, field in enumerate(field_names, start=1):
            value = analysis.get(field, "")

            if field == "best_fit_states" and isinstance(value, list):
                value = "; ".join(value)

            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

            if row_idx % 2 == 0:
                cell.fill = data_fill

    # -----------------------------
    # COLUMN SIZING
    # -----------------------------
    for col in range(1, len(field_names) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20

    ws.row_dimensions[1].height = 50

    wb.save(filename)
    return filename
def append_analyses_to_excel(analyses, filename=r'C:\Users\HP\Downloads\Therapy_Video_Scoring_Sheet_With_Dropdowns.xlsx'):
    """
    Appends new analysis rows to an EXISTING Excel file that already has:
    - headers in row 1
    - data validations / dropdowns already set up
    """
    if not analyses:
        print("No new analyses to append.")
        return filename

    try:
        wb = openpyxl.load_workbook(filename, data_only=False)  # keep formulas & formatting
        ws = wb['Therapy Video Scoring']                       # ← your sheet name
    except FileNotFoundError:
        print(f"File not found: {filename}")
        return None
    except KeyError:
        print(f"Sheet 'Therapy Video Scoring' not found in {filename}")
        return None

    # Find the next empty row (look at column A – video_url)
    next_row = 2
    while ws.cell(row=next_row, column=1).value is not None:
        next_row += 1

    print(f"Appending {len(analyses)} row(s) starting from row {next_row}")

    # Your field order — MUST match the column order in the existing Excel
    field_names = [
        "video_url", "language", "primary_modality", "language_presence",
        "therapeutic_explicitness", "session_structure", "pacing_adequacy",
        "source_lineage", "symbolic_density", "modern_intrusion_level",
        "emotional_valence", "affective_stability", "arousal_trajectory",
        "instruction_density", "cognitive_demand", "inner_silence_support",
        "directive_strength", "dependency_risk", "claim_severity",
        "autonomic_orientation", "best_fit_states", "sleep_proximity_safety",
        "cultural_accessibility", "locale_fit", "environmental_compatibility",
        "sensory_purity", "rhythmic_stability", "visual_motion_profile",
        "repetitiveness_quality", "attention_capture_style", "internal_consistency",
        "production_intrusion", "commercial_influence",
        "therapeutic_suitability_score", "Risk Flag Summary", "One-line Neutral Description"
    ]

    # Write the new rows
    for analysis in analyses:
        for col_idx, field in enumerate(field_names, start=1):
            value = analysis.get(field, "")

            if field == "best_fit_states" and isinstance(value, list):
                value = "; ".join(str(x).strip() for x in value if x)

            cell = ws.cell(row=next_row, column=col_idx, value=value)
            # Optional: light formatting for new rows
            cell.alignment = openpyxl.styles.Alignment(horizontal="left", vertical="center", wrap_text=True)

        next_row += 1

    try:
        wb.save(filename)
        print(f"Successfully appended data to: {filename}")
        print(f"Last row now ≈ {next_row-1}")
    except PermissionError:
        print("✘ Could not save — file is open in Excel? Close it and try again.")
    except Exception as e:
        print(f"Error while saving: {e}")

    return filename

def main():
    """Main execution function."""
    print("=" * 80)
    print("YouTube Video Wellness Analysis - Excel Export Tool")
    print("=" * 80)
    print("\nThis tool analyzes videos and exports structured features to Excel")
    print("All features are assigned values from predefined options only\n")
    
    if not HAS_YT_DLP_MODULE:
        print("⚠️  Installing yt-dlp...")
        import subprocess
        subprocess.run([sys.executable, '-m', 'pip', 'install', 'yt-dlp', '-q'])
        print("Installed. Please run again.\n")
        return
    
    analyses = []
    
    while True:
        url = input("\nEnter YouTube URL (or 'done' to export to Excel): ").strip()
        
        if url.lower() == 'done':
            break
        
        if not url:
            continue
        
        video_id = extract_video_id(url)
        if not video_id:
            print("❌ Invalid YouTube URL format")
            continue
        
        print(f"\n[1/3] Extracting metadata...")
        metadata = get_video_metadata(video_id)
        if metadata:
            print(f"  ✓ Title: {metadata.get('title', 'Unknown')[:60]}")
        
        print(f"[2/3] Extracting transcript...")
        transcript = None
        audio_path = None
        duration = 0
        
        # Download audio and get transcript
        with download_audio_temp(url) as (audio_path, duration):
            transcript = get_transcript(video_id, audio_path, duration)
            if transcript:
                print(f"  ✓ Extracted {len(transcript.split())} words")
            else:
                print(f"  ⚠️  No captions available (proceeding with metadata)")
            
            print(f"[3/3] Analyzing with Groq...")
            analysis = analyze_video_with_groq(url, transcript, metadata)
            if 'error' not in analysis:
                # NEW: Audio-based pacing analysis
                audio_pacing = analyze_audio_pacing(audio_path, duration)
                if audio_pacing:
                    print(f"  ✓ Audio-calculated pacing: {audio_pacing} (overriding LLM)")
                    analysis["pacing_adequacy"] = audio_pacing

                # Compute component scores and final rounded score
                component_scores = {}
                total = 0.0
                count = 0

                for field, mapping in SCORE_MAP.items():
                    value = analysis.get(field)
                    if value in mapping:
                        score = mapping[value]
                        component_scores[field] = score
                        total += score
                        count += 1

                if count > 0:
                    average = total / count
                    final_score = round(average)           # nearest integer
                    # or: final_score = round(average * 2) / 2  ← if you want 0.5 steps
                    analysis["therapeutic_suitability_score"] = str(final_score)
                    print("average",average)
                    # Optional: store for debugging
                    analysis["_component_scores"] = component_scores
                    analysis["_raw_average"] = round(average, 2)
                    print(f"  ✓ Analysis complete - Suitability Score: {final_score}/5")
                else:
                    analysis["therapeutic_suitability_score"] = "N/A"
        
        if 'error' in analysis:
            print(f"  ❌ Error: {analysis['error']}")
        else:
            score = analysis.get('therapeutic_suitability_score', 'N/A')
            print(f"  ✓ Analysis complete - Suitability Score: {score}/5")
            analyses.append(analysis)
    
    if analyses:
        print("\n" + "=" * 80)
        target_file = r'C:\Users\HP\Downloads\Therapy_Video_Scoring_Sheet_With_Dropdowns.xlsx'
        print(f"Attempting to append {len(analyses)} video(s) to:")
        print(f"  {target_file}")
        print("-" * 70)
        
        result = append_analyses_to_excel(analyses, target_file)
        
        if result:
            print(f"✓ Done. File updated successfully.")
        else:
            print("✘ Append failed.")
        print("=" * 80)
    else:
        print("\nNo new analyses to append.")

def read_links_from_therapy_tracker(
    filename=r'C:\Users\HP\Downloads\TherapyTracker.xlsx',
    sheet_name='Indian',
    column_name='Audio/Video Resource'
):
    """Read YouTube links from TherapyTracker Excel file."""
    try:
        wb = openpyxl.load_workbook(filename, data_only=True)
        ws = wb[sheet_name]
        
        # Find the column index by header name
        header_row = 1
        column_idx = None
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col).value
            if cell_value and column_name.lower() in str(cell_value).lower():
                column_idx = col
                break
        
        if column_idx is None:
            print(f"❌ Column '{column_name}' not found in sheet '{sheet_name}'")
            return []
        
        print(f"✓ Found column '{column_name}' at column {column_idx}")
        
        # Extract all URLs from that column (skip header)
        urls = []
        for row in range(2, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=column_idx).value
            if cell_value:
                cell_str = str(cell_value).strip()
                if 'youtube.com' in cell_str or 'youtu.be' in cell_str:
                    urls.append(cell_str)
        
        print(f"✓ Found {len(urls)} YouTube links in the sheet")
        return urls
    
    except FileNotFoundError:
        print(f"❌ File not found: {filename}")
        return []
    except KeyError:
        print(f"❌ Sheet '{sheet_name}' not found in {filename}")
        return []
    except Exception as e:
        print(f"❌ Error reading Excel: {str(e)}")
        return []

def process_single_video(url):
    """Process a single video URL and return analysis dict."""
    video_id = extract_video_id(url)
    if not video_id:
        print(f"  ❌ Invalid YouTube URL format")
        return None
    
    print(f"  [1/3] Extracting metadata...")
    metadata = get_video_metadata(video_id)
    if metadata:
        print(f"    ✓ Title: {metadata.get('title', 'Unknown')[:50]}")
    
    # Download audio and get transcript
    with download_audio_temp(url) as (audio_path, duration):
        print(f"  [2/3] Extracting transcript...")
        transcript = get_transcript(video_id, audio_path, duration)
        if transcript:
            print(f"    ✓ Extracted {len(transcript.split())} words")
        else:
            print(f"    ⚠️  No captions available")
        
        print(f"  [3/3] Analyzing with Groq...")
        analysis = analyze_video_with_groq(url, transcript, metadata)
        
        if 'error' not in analysis:
            # NEW: Audio-based pacing analysis
            audio_pacing = analyze_audio_pacing(audio_path, duration)
            if audio_pacing:
                print(f"    ✓ Audio-calculated pacing: {audio_pacing} (overriding LLM)")
                analysis["pacing_adequacy"] = audio_pacing
            
            component_scores = {}
            total = 0.0
            count = 0

            for field, mapping in SCORE_MAP.items():
                value = analysis.get(field)
                if value in mapping:
                    score = mapping[value]
                    component_scores[field] = score
                    total += score
                    count += 1

            if count > 0:
                average = total / count
                final_score = round(average)
                analysis["therapeutic_suitability_score"] = str(final_score)
                print(f"    ✓ Score: {final_score}/5")
            else:
                analysis["therapeutic_suitability_score"] = "N/A"
            return analysis
        else:
            print(f"    ❌ Error: {analysis['error']}")
            return None

def batch_process():
    """Batch process all links from TherapyTracker.xlsx using parallel execution."""
    print("=" * 80)
    print("BATCH MODE: Processing links from TherapyTracker.xlsx")
    print("=" * 80)
    
    urls = read_links_from_therapy_tracker()
    
    if not urls:
        print("\n❌ No YouTube links found.")
        return
    
    # Limit to first 100 links (or user specified slice)
    urls = urls[15:30]
    
    print(f"\nProcessing {len(urls)} videos with parallel execution (max_workers=3)...\n")
    
    analyses = []
    skipped = 0
    completed = 0
    target_file = r'C:\Users\HP\Downloads\Therapy_Video_Scoring_Sheet_With_Dropdowns.xlsx'
    
    from concurrent.futures import ThreadPoolExecutor, as_completed
    
    # Use ThreadPoolExecutor for parallel processing
    with ThreadPoolExecutor(max_workers=3) as executor:
        # Submit all tasks
        future_to_url = {executor.submit(process_single_video, url): url for url in urls}
        
        for i, future in enumerate(as_completed(future_to_url), 1):
            url = future_to_url[future]
            try:
                analysis = future.result()
                if analysis:
                    analyses.append(analysis)
                    completed += 1
                else:
                    skipped += 1
            except Exception as e:
                print(f"  ❌ Error processing {url[:30]}...: {str(e)}")
                skipped += 1
            
            # Print progress
            print(f"  [Progress] {i}/{len(urls)} completed")
            
            # Auto-save every 5 COMPLETED videos
            if len(analyses) >= 5:
                print(f"\n💾 Batch save: appending {len(analyses)} videos...")
                append_analyses_to_excel(analyses, target_file)
                analyses = []  # Clear buffer after save

    # Final save of any remaining analyses
    if analyses:
        print(f"\n💾 Final save: appending {len(analyses)} videos...")
        append_analyses_to_excel(analyses, target_file)
    
    print("\n" + "=" * 80)
    print(f"BATCH PROCESSING COMPLETE")
    print(f"  ✓ Processed: {completed}")
    print(f"  ⚠️  Skipped/Failed: {skipped}")
    print("=" * 80)

if __name__ == '__main__':
    import sys
    if len(sys.argv) > 1 and sys.argv[1] == '--batch':
        batch_process()
    else:
        print("Run with --batch flag to process TherapyTracker.xlsx")
        print("  python bwsscript.py --batch")
        print("\nOr run without flag for interactive mode:")
        main()
